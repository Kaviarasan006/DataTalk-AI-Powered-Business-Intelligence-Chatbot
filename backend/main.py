from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
import pandas as pd
import json
import os
import uuid
import io
import re
import requests
import plotly.express as px
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

app = FastAPI(title="DataTalk API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "/tmp/datatalk_uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

sessions: dict = {}


def call_gemini(prompt: str) -> str:
    api_key = os.getenv("GROQ_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="GROQ_API_KEY not set. Add it to your .env file.")

    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": "llama-3.3-70b-versatile",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 1024,
    }
    resp = requests.post(url, json=payload, headers=headers, timeout=30)
    if resp.status_code != 200:
        raise HTTPException(status_code=500, detail=f"Groq API error: {resp.text[:300]}")
    data = resp.json()
    return data["choices"][0]["message"]["content"]


class QueryRequest(BaseModel):
    session_id: str
    question: str


class ExportRequest(BaseModel):
    session_id: str


@app.get("/")
def root():
    return {"status": "DataTalk API running", "version": "1.0.0"}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    session_id = str(uuid.uuid4())
    contents = await file.read()

    try:
        if file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents))
        elif file.filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files supported.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {str(e)}")

    df.columns = [re.sub(r"[^a-zA-Z0-9_]", "_", c.strip().lower()) for c in df.columns]

    file_path = os.path.join(UPLOAD_DIR, f"{session_id}.parquet")
    df.to_parquet(file_path, index=False)

    schema = {}
    for col in df.columns:
        dtype = str(df[col].dtype)
        sample = df[col].dropna().head(3).tolist()
        schema[col] = {"type": dtype, "sample": [str(s) for s in sample]}

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    stats = {}
    if numeric_cols:
        stats = df[numeric_cols].describe().round(2).to_dict()

    sessions[session_id] = {
        "file_path": file_path,
        "filename": file.filename,
        "schema": schema,
        "columns": df.columns.tolist(),
        "row_count": len(df),
        "history": [],
    }

    return {
        "session_id": session_id,
        "filename": file.filename,
        "rows": len(df),
        "columns": df.columns.tolist(),
        "schema": schema,
        "stats": stats,
    }


@app.post("/query")
async def query_data(req: QueryRequest):
    if req.session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found. Please upload a file first.")

    session = sessions[req.session_id]
    df = pd.read_parquet(session["file_path"])
    schema_str = json.dumps(session["schema"], indent=2)
    columns = session["columns"]

    history_str = ""
    for h in session["history"][-4:]:
        history_str += f"User: {h['question']}\nSummary: {h['summary']}\n\n"

    prompt = f"""You are DataTalk, an expert data analyst AI. The user uploaded a dataset.

DATASET INFO:
- Columns: {columns}
- Row count: {session['row_count']}
- Schema (column name → type + sample values):
{schema_str}

RECENT CONVERSATION:
{history_str}

USER QUESTION: {req.question}

TASK: Write a pandas expression to answer the question, pick a chart type, and write a plain-English insight.

RULES:
- Use ONLY column names listed above (they are already lowercase with underscores)
- Write ONE pandas expression assigned to `result`
- `result` must be a DataFrame or Series
- Use .reset_index() after groupby
- Always .head(20) for large results
- For date/month analysis, use: df['date'] = pd.to_datetime(df['date']); then df.groupby(df['date'].dt.to_period('M'))...

Respond ONLY with valid JSON — no markdown, no backticks, no extra text:
{{"code":"result = df.groupby('category')['revenue'].sum().reset_index().sort_values('revenue',ascending=False).head(10)","chart_type":"bar","x_col":"category","y_col":"revenue","chart_title":"Revenue by Category","insight":"Electronics leads with the highest revenue at ₹X, followed by Clothing.","summary":"Grouped revenue by category"}}
"""

    raw = call_gemini(prompt)

    # Strip markdown fences if present
    raw = re.sub(r"```json\s*", "", raw).strip()
    raw = re.sub(r"```\s*", "", raw).strip()

    try:
        parsed = json.loads(raw)
    except Exception:
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if match:
            try:
                parsed = json.loads(match.group())
            except Exception:
                raise HTTPException(status_code=500, detail=f"Could not parse LLM response: {raw[:300]}")
        else:
            raise HTTPException(status_code=500, detail=f"LLM returned unexpected format: {raw[:300]}")

    code = parsed.get("code", "")
    chart_type = parsed.get("chart_type", "none")
    x_col = parsed.get("x_col", "")
    y_col = parsed.get("y_col", "")
    chart_title = parsed.get("chart_title", req.question)
    insight = parsed.get("insight", "")
    summary = parsed.get("summary", "")

    local_vars = {"df": df.copy(), "pd": pd, "result": None}
    try:
        exec(code, {"pd": pd, "__builtins__": {}}, local_vars)
        result = local_vars.get("result")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Query execution error: {str(e)}\nGenerated code: {code}")

    if result is None:
        raise HTTPException(status_code=422, detail="Query returned no result.")

    if isinstance(result, pd.Series):
        result = result.reset_index()
        result.columns = [str(c) for c in result.columns]

    result_df = result if isinstance(result, pd.DataFrame) else pd.DataFrame({"value": [result]})
    result_df = result_df.head(50)

    # Convert Period columns to string for JSON serialization
    for col in result_df.columns:
        if hasattr(result_df[col], 'dt'):
            result_df[col] = result_df[col].astype(str)
        elif result_df[col].dtype == object:
            result_df[col] = result_df[col].astype(str)

    chart_json = None
    colors = ["#6366f1","#8b5cf6","#06b6d4","#10b981","#f59e0b","#ef4444","#ec4899","#14b8a6"]

    if chart_type != "none" and x_col in result_df.columns and y_col in result_df.columns:
        try:
            if chart_type == "bar":
                fig = px.bar(result_df, x=x_col, y=y_col, title=chart_title, color_discrete_sequence=colors)
            elif chart_type == "line":
                fig = px.line(result_df, x=x_col, y=y_col, title=chart_title, color_discrete_sequence=colors)
            elif chart_type == "pie":
                fig = px.pie(result_df, names=x_col, values=y_col, title=chart_title, color_discrete_sequence=colors)
            elif chart_type == "scatter":
                fig = px.scatter(result_df, x=x_col, y=y_col, title=chart_title, color_discrete_sequence=colors)
            elif chart_type == "histogram":
                fig = px.histogram(result_df, x=x_col, title=chart_title, color_discrete_sequence=colors)
            else:
                fig = None

            if fig:
                fig.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Inter, sans-serif", size=13, color="#94a3b8"),
                    title_font=dict(size=16, color="#e2e8f0"),
                    margin=dict(l=40, r=20, t=50, b=40),
                )
                chart_json = fig.to_json()
        except Exception as e:
            chart_json = None

    table_data = result_df.to_dict(orient="records")
    table_cols = result_df.columns.tolist()

    session["history"].append({
        "question": req.question,
        "code": code,
        "summary": summary,
        "insight": insight,
    })

    return {
        "insight": insight,
        "chart": chart_json,
        "chart_type": chart_type,
        "chart_title": chart_title,
        "table": table_data,
        "columns": table_cols,
        "code": code,
        "row_count": len(result_df),
    }


@app.post("/export/excel")
async def export_excel(req: ExportRequest):
    if req.session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found.")

    session = sessions[req.session_id]
    df = pd.read_parquet(session["file_path"])
    history = session["history"]

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1e1b4b")
    accent_fill = PatternFill("solid", fgColor="4f46e5")
    light_fill = PatternFill("solid", fgColor="eef2ff")

    for col in ["A","B","C","D","E","F","G","H"]:
        ws_summary.column_dimensions[col].width = 22

    ws_summary.row_dimensions[2].height = 40
    ws_summary["B2"] = "DataTalk — AI Business Intelligence Report"
    ws_summary["B2"].font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    ws_summary["B2"].fill = header_fill
    ws_summary["B2"].alignment = Alignment(vertical="center")
    ws_summary.merge_cells("B2:H2")

    meta = [("Dataset", session["filename"]), ("Total Rows", session["row_count"]),
            ("Columns", len(session["columns"])), ("Questions Asked", len(history))]
    for i, (k, v) in enumerate(meta, 4):
        ws_summary[f"B{i}"] = k
        ws_summary[f"B{i}"].font = Font(name="Calibri", bold=True, size=11)
        ws_summary[f"C{i}"] = v
        ws_summary[f"C{i}"].font = Font(name="Calibri", size=11)

    ws_summary["B9"] = "AI-Generated Insights"
    ws_summary["B9"].font = Font(name="Calibri", bold=True, size=13, color="4f46e5")

    for i, h in enumerate(history):
        row = 11 + i * 3
        ws_summary[f"B{row}"] = f"Q{i+1}: {h['question']}"
        ws_summary[f"B{row}"].font = Font(name="Calibri", bold=True, size=11)
        ws_summary[f"B{row}"].fill = light_fill
        ws_summary.merge_cells(f"B{row}:H{row}")
        ws_summary[f"B{row+1}"] = h.get("insight", "")
        ws_summary[f"B{row+1}"].font = Font(name="Calibri", size=10, italic=True)
        ws_summary[f"B{row+1}"].alignment = Alignment(wrap_text=True)
        ws_summary.merge_cells(f"B{row+1}:H{row+1}")
        ws_summary.row_dimensions[row+1].height = 30

    # Raw data sheet
    ws_data = wb.create_sheet("Raw Data")
    ws_data.sheet_view.showGridLines = False
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center")
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_name) + 4)
    for row_idx, row in enumerate(df.head(500).itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            if row_idx % 2 == 0:
                cell.fill = light_fill

    # Stats sheet
    numeric_df = df.select_dtypes(include="number")
    if not numeric_df.empty:
        ws_stats = wb.create_sheet("Statistics")
        ws_stats.sheet_view.showGridLines = False
        stats_df = numeric_df.describe().round(2).reset_index()
        for col_idx, col_name in enumerate(stats_df.columns, 2):
            cell = ws_stats.cell(row=2, column=col_idx, value=str(col_name).replace("_", " ").title())
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = accent_fill
            cell.alignment = Alignment(horizontal="center")
            ws_stats.column_dimensions[get_column_letter(col_idx)].width = 18
        for row_idx, row in enumerate(stats_df.itertuples(index=False), 3):
            for col_idx, val in enumerate(row, 2):
                cell = ws_stats.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                if row_idx % 2 == 0:
                    cell.fill = light_fill

    out_path = f"/tmp/datatalk_report_{req.session_id}.xlsx"
    wb.save(out_path)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"DataTalk_Report_{session['filename'].split('.')[0]}.xlsx",
    )


@app.get("/session/{session_id}")
def get_session(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found.")
    s = sessions[session_id]
    return {"filename": s["filename"], "rows": s["row_count"],
            "columns": s["columns"], "schema": s["schema"], "history": s["history"]}